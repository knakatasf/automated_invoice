import openpyxl
from openpyxl import load_workbook
from openpyxl import drawing
from openpyxl.styles import PatternFill

def makeFile(templatePath, outputPath):
    workbook = openpyxl.load_workbook(templatePath)
    workbook.save(outputPath)

def pasteLogo(sheet):
    # Make image object
    logo = drawing.image.Image("/Users/lorh/Desktop/Automated_Invoice/lohkLogo.png")
    logo.width = 850
    logo.height = 135
    sheet.add_image(logo, "A1") # Insert image object (size already adjusted)


def fillBaseInfo(sheet, baseInfoList):
    # Matter is stored in the last and fill in cell C8
    matterStr = baseInfoList.pop()
    sheet.cell(row=8, column=3, value=matterStr)

    for rowNum, value in enumerate(baseInfoList, start=7):
        sheet.cell(row=rowNum, column=5, value=value)

def findEndRow(sheet, rowNum):
    columnB = sheet.iter_cols(min_col=2, max_col=2, # Only qualify column B
                              min_row=rowNum, values_only=True)
    
    for column in columnB:
        for value in column:
            if value is not None:
                return rowNum
            rowNum += 1

    return None

def fillFormula(sheet, rowNum): # rowNum is the row of "Attorney Services Rendered:"
    rowNum += 1                 # Now rowNum is the row of "Managing Partner"
    # Managing Partner HOURS and AMOUNT
    sheet[f"D{rowNum}"] = f"=SUMIF(E15:E{rowNum-1}, \">=500\", D15:D{rowNum-1})"
    sheet[f"F{rowNum}"] = f"=SUMIF(E15:E{rowNum-1}, \">=500\", F15:F{rowNum-1})"

    # Partner HOURS and AMOUNT
    sheet[f"D{rowNum+1}"] = f"=SUMIFS(D15:D{rowNum-1}, E15:E{rowNum-1}, \">=400\", E15:E{rowNum-1}, \"<500\")"
    sheet[f"F{rowNum+1}"] = f"=SUMIFS(F15:F{rowNum-1}, E15:E{rowNum-1}, \">=400\", E15:E{rowNum-1}, \"<500\")"

    # Associate Attorney HOURS and AMOUNT
    sheet[f"D{rowNum+2}"] = f"=SUMIFS(D15:D{rowNum-1}, E15:E{rowNum-1}, \">=250\", E15:E{rowNum-1}, \"<400\")"
    sheet[f"F{rowNum+2}"] = f"=SUMIFS(F15:F{rowNum-1}, E15:E{rowNum-1}, \">=250\", E15:E{rowNum-1}, \"<400\")"

    # Law Clark/Paralegal HOURS and AMOUNT
    sheet[f"D{rowNum+3}"] = f"=SUMIFS(D15:D{rowNum-1}, E15:E{rowNum-1}, \">=200\", E15:E{rowNum-1}, \"<250\")"
    sheet[f"F{rowNum+3}"] = f"=SUMIFS(F15:F{rowNum-1}, E15:E{rowNum-1}, \">=200\", E15:E{rowNum-1}, \"<250\")"

    # Legal Assistant HOURS and AMOUNT
    sheet[f"D{rowNum+4}"] = f"=SUMIF(E15:E{rowNum-1}, \"<200\", D15:D{rowNum-1})"
    sheet[f"F{rowNum+4}"] = f"=SUMIF(E15:E{rowNum-1}, \"<200\", F15:F{rowNum-1})"

    # Total Attorneys' Fees:
    sheet[f"F{rowNum+5}"] = f"=SUM(F{rowNum}:F{rowNum+4})"

    # TOTAL AMOUNT CURRENTLY DUE:
    sheet[f"F{rowNum+13}"] = f"=SUM(F{rowNum+5}, F{rowNum+8}:F{rowNum+8}, F{rowNum+11}:F{rowNum+11})"

def makeInvoice(outputPath, baseInfoDict, dataDict, invDate):
    workbook = load_workbook(outputPath)
    templateSheet = workbook.active

    for client, entryList in dataDict.items():
        newSheet = workbook.copy_worksheet(templateSheet)

        pasteLogo(newSheet)

        # Name sheet with client name as appears in ACR
        newSheet.title = client

        # Fill invoice issue date
        newSheet.cell(row=7, column=3, value=invDate)

        # baseInfo is Bill To and Matter section. Matter is the last in the list.
        baseInfoList = baseInfoDict[client]
        fillBaseInfo(newSheet, baseInfoList)

        lightGrey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        rowNum = 15 # Start row. Need to specify!
        for rowData in entryList:
            for colNum, value in enumerate(rowData, 2): # Start column. Need to specify!
                cell = newSheet.cell(row=rowNum, column=colNum, value=value)
                if rowNum%2 == 0: # If row # is even, fill lightgrey in cells
                    cell.fill = lightGrey
            
            timeValue = newSheet[f"D{rowNum}"].value
            if timeValue != "Flat Fee" and timeValue != "Not Billed":
                cell = newSheet.cell(row=rowNum, column=6, value=f"=D{rowNum}*E{rowNum}")
            else:
                cell = newSheet.cell(row=rowNum, column=6, value=0)
            if rowNum%2 == 0:
                cell.fill = lightGrey

            rowNum += 1
        
        endRow = findEndRow(newSheet, rowNum) # Want to find the first row which is not empty
        numToDelete = endRow - rowNum

        for _ in range(numToDelete):
             newSheet.delete_rows(rowNum)

        fillFormula(newSheet, rowNum)


    workbook.save(outputPath)
